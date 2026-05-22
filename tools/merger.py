"""
merger.py — Merge multiple Excel files or sheets into one master workbook.

Supports two modes:
  1. Multi-file merge: combine uploaded .xlsx files (each becomes a sheet, or all stacked)
  2. Multi-sheet stack: read every sheet from a single workbook and stack rows vertically
"""

import io
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter


def merge_files(uploaded_files: list, mode: str = "stack") -> bytes:
    """
    Merge multiple uploaded Excel files.

    Args:
        uploaded_files: list of file-like objects (Streamlit UploadedFile)
        mode:
            "stack"  — append all sheets row-by-row into one sheet called 'Master'
            "sheets" — keep each file as its own sheet, named after the file

    Returns:
        bytes of the resulting .xlsx file
    """
    frames = []
    sheet_map = {}  # sheet_name -> DataFrame  (used for "sheets" mode)

    for f in uploaded_files:
        raw = f.read()
        file_name = f.name.replace(".xlsx", "").replace(".xls", "")
        xls = pd.ExcelFile(io.BytesIO(raw))

        for sheet in xls.sheet_names:
            df = pd.read_excel(io.BytesIO(raw), sheet_name=sheet, dtype=str)
            df["_source_file"] = file_name
            df["_source_sheet"] = sheet

            if mode == "stack":
                frames.append(df)
            else:
                label = f"{file_name} – {sheet}" if len(xls.sheet_names) > 1 else file_name
                label = label[:31]  # Excel sheet name limit
                sheet_map[label] = df

    output = io.BytesIO()

    if mode == "stack":
        master = pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()
        with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
            master.to_excel(writer, sheet_name="Master", index=False)
            _style_header(writer, "Master", master)
    else:
        with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
            for sheet_name, df in sheet_map.items():
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                _style_header(writer, sheet_name, df)

    return output.getvalue()


def merge_sheets(uploaded_file) -> bytes:
    """
    Read every sheet from a single workbook and stack them vertically into one 'Master' sheet.
    A '_sheet' column is added to track the origin.

    Args:
        uploaded_file: file-like object

    Returns:
        bytes of the resulting .xlsx file
    """
    raw = uploaded_file.read()
    xls = pd.ExcelFile(io.BytesIO(raw))
    frames = []

    for sheet in xls.sheet_names:
        df = pd.read_excel(io.BytesIO(raw), sheet_name=sheet, dtype=str)
        df.insert(0, "_sheet", sheet)
        frames.append(df)

    master = pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
        master.to_excel(writer, sheet_name="Master", index=False)
        _style_header(writer, "Master", master)

    return output.getvalue()


def _style_header(writer, sheet_name: str, df: pd.DataFrame):
    """Apply bold header row + auto column width."""
    workbook = writer.book
    worksheet = writer.sheets[sheet_name]

    header_fmt = workbook.add_format({
        "bold": True,
        "bg_color": "#1F4E79",
        "font_color": "#FFFFFF",
        "border": 1,
        "align": "center",
        "valign": "vcenter",
    })

    for col_num, col_name in enumerate(df.columns):
        worksheet.write(0, col_num, col_name, header_fmt)
        col_width = max(len(str(col_name)) + 4, 12)
        worksheet.set_column(col_num, col_num, col_width)

    worksheet.freeze_panes(1, 0)
