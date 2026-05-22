"""
formatter.py — Auto-format an Excel file and highlight exceptions.

Rules applied:
  • Overdue dates   — any date column where the date < today → red fill
  • Budget overruns — any numeric column whose header contains "budget", "cost",
                      "spend", or "amount" and value exceeds a user-supplied cap → orange fill
  • Negative values — any negative number → yellow fill
  • Header row      — dark blue background, white bold text
  • Auto column widths + freeze top row
"""

import io
from datetime import date, datetime

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── Colour palette ─────────────────────────────────────────────────────────────
HEADER_FILL   = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT   = Font(bold=True, color="FFFFFF")
OVERDUE_FILL  = PatternFill("solid", fgColor="FF4C4C")   # red
OVERRUN_FILL  = PatternFill("solid", fgColor="FF9900")   # orange
NEGATIVE_FILL = PatternFill("solid", fgColor="FFEB9C")   # yellow
THIN_BORDER   = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)

_DATE_KEYWORDS    = {"date", "due", "deadline", "expiry", "expiration", "start", "end"}
_BUDGET_KEYWORDS  = {"budget", "cost", "spend", "amount", "expense", "actual", "forecast"}


def format_and_highlight(
    uploaded_file,
    overrun_threshold: float = 0.0,
    flag_overdue: bool = True,
    flag_negatives: bool = True,
) -> bytes:
    """
    Load an Excel file, apply formatting and exception highlighting, return bytes.

    Args:
        uploaded_file:      file-like object (.xlsx)
        overrun_threshold:  flag numeric cells above this value (0 = disabled)
        flag_overdue:       highlight date cells that are in the past
        flag_negatives:     highlight negative numeric cells

    Returns:
        bytes of the formatted .xlsx file
    """
    raw = uploaded_file.read()
    df = pd.read_excel(io.BytesIO(raw))

    # Write the clean DataFrame to a new workbook via openpyxl
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Formatted")

    output.seek(0)
    wb = load_workbook(output)
    ws = wb.active

    today = date.today()
    max_col = ws.max_column
    max_row = ws.max_row

    # Collect column metadata from header row
    col_meta = {}  # col_index (1-based) -> {"name": str, "is_date": bool, "is_budget": bool}
    for col_idx in range(1, max_col + 1):
        header_cell = ws.cell(row=1, column=col_idx)
        col_name = str(header_cell.value or "").lower()
        col_meta[col_idx] = {
            "name": col_name,
            "is_date":   any(k in col_name for k in _DATE_KEYWORDS),
            "is_budget": any(k in col_name for k in _BUDGET_KEYWORDS),
        }

    # ── Style header row ───────────────────────────────────────────────────────
    for col_idx in range(1, max_col + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill   = HEADER_FILL
        cell.font   = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # ── Process data rows ──────────────────────────────────────────────────────
    for row_idx in range(2, max_row + 1):
        for col_idx in range(1, max_col + 1):
            cell  = ws.cell(row=row_idx, column=col_idx)
            meta  = col_meta[col_idx]
            value = cell.value
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center")

            if value is None:
                continue

            # Overdue date check
            if flag_overdue and meta["is_date"]:
                cell_date = _to_date(value)
                if cell_date and cell_date < today:
                    cell.fill = OVERDUE_FILL
                    continue

            # Budget overrun check
            if overrun_threshold > 0 and meta["is_budget"]:
                try:
                    if float(value) > overrun_threshold:
                        cell.fill = OVERRUN_FILL
                        continue
                except (ValueError, TypeError):
                    pass

            # Negative value check
            if flag_negatives:
                try:
                    if float(value) < 0:
                        cell.fill = NEGATIVE_FILL
                        continue
                except (ValueError, TypeError):
                    pass

    # ── Auto column widths ─────────────────────────────────────────────────────
    for col_idx in range(1, max_col + 1):
        col_letter = get_column_letter(col_idx)
        max_width  = 10
        for row_idx in range(1, min(max_row + 1, 200)):  # sample first 200 rows
            val = ws.cell(row=row_idx, column=col_idx).value
            if val:
                max_width = max(max_width, len(str(val)) + 4)
        ws.column_dimensions[col_letter].width = min(max_width, 50)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    result = io.BytesIO()
    wb.save(result)
    return result.getvalue()


def _to_date(value) -> date | None:
    """Attempt to coerce a cell value to a date object."""
    if isinstance(value, (datetime,)):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y", "%d-%b-%Y"):
            try:
                return datetime.strptime(value.strip(), fmt).date()
            except ValueError:
                continue
    return None
