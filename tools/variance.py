"""
variance.py — Budget Variance Reporter

Expects an Excel file with at minimum:
  • A "Category" (or "Department" / "Item") column
  • A "Budget" (or "Forecast") column
  • An "Actual" column

Produces a new workbook with:
  • Original data + Variance ($) + Variance (%) columns
  • Conditional colour coding:
      - Green  → under budget (actual < budget)
      - Orange → variance 0–10 %
      - Red    → variance > threshold (default 10 %)
  • A pivot-style summary table on a second sheet
  • A bar chart comparing Budget vs Actual per category
"""

import io
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabel


# ── Colours ────────────────────────────────────────────────────────────────────
HEADER_FILL  = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT  = Font(bold=True, color="FFFFFF")
GREEN_FILL   = PatternFill("solid", fgColor="C6EFCE")   # under budget
ORANGE_FILL  = PatternFill("solid", fgColor="FFEB9C")   # 0-threshold %
RED_FILL     = PatternFill("solid", fgColor="FFC7CE")   # over threshold
SUBTOTAL_FILL= PatternFill("solid", fgColor="D9E1F2")   # summary rows
THIN = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)
MEDIUM_BOTTOM = Border(bottom=Side(style="medium"))

_CATEGORY_ALIASES = ["category", "department", "item", "description", "name", "account"]
_BUDGET_ALIASES   = ["budget", "forecast", "plan", "planned"]
_ACTUAL_ALIASES   = ["actual", "actuals", "spend", "spent", "real"]


def generate_variance_report(
    uploaded_file,
    variance_threshold: float = 10.0,
) -> bytes:
    """
    Build a variance report workbook.

    Args:
        uploaded_file:       file-like object (.xlsx)
        variance_threshold:  percentage above which a row is flagged red (default 10 %)

    Returns:
        bytes of the resulting .xlsx
    """
    raw = uploaded_file.read()
    df = pd.read_excel(io.BytesIO(raw))

    # ── Column detection ───────────────────────────────────────────────────────
    col_map = _detect_columns(df)
    if not col_map:
        raise ValueError(
            "Could not find Budget and Actual columns. "
            "Please ensure your file has columns named Budget/Forecast and Actual."
        )

    cat_col    = col_map.get("category")
    budget_col = col_map["budget"]
    actual_col = col_map["actual"]

    # ── Compute variance ───────────────────────────────────────────────────────
    df[budget_col] = pd.to_numeric(df[budget_col], errors="coerce").fillna(0)
    df[actual_col] = pd.to_numeric(df[actual_col], errors="coerce").fillna(0)

    df["Variance ($)"] = df[actual_col] - df[budget_col]
    df["Variance (%)"] = df.apply(
        lambda r: (r["Variance ($)"] / r[budget_col] * 100) if r[budget_col] != 0 else 0,
        axis=1,
    ).round(2)
    df["Status"] = df["Variance (%)"].apply(
        lambda v: "Over Budget" if v > variance_threshold
        else ("On Track" if v >= 0 else "Under Budget")
    )

    # ── Write to openpyxl ──────────────────────────────────────────────────────
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Variance Report")

        # Build summary sheet
        if cat_col:
            summary = (
                df.groupby(cat_col)[[budget_col, actual_col, "Variance ($)"]].sum()
                .reset_index()
            )
            summary["Variance (%)"] = summary.apply(
                lambda r: round(r["Variance ($)"] / r[budget_col] * 100, 2) if r[budget_col] else 0,
                axis=1,
            )
            summary.to_excel(writer, index=False, sheet_name="Summary")

    output.seek(0)
    wb = load_workbook(output)

    # ── Style "Variance Report" sheet ──────────────────────────────────────────
    ws = wb["Variance Report"]
    _style_header_row(ws, df)

    col_letters = {name: get_column_letter(i + 1) for i, name in enumerate(df.columns)}
    var_pct_col = col_letters.get("Variance (%)")
    status_col  = col_letters.get("Status")

    for row_idx in range(2, ws.max_row + 1):
        try:
            pct = float(ws[f"{var_pct_col}{row_idx}"].value or 0)
        except (ValueError, TypeError):
            pct = 0

        fill = (
            RED_FILL    if pct > variance_threshold else
            ORANGE_FILL if pct >= 0                 else
            GREEN_FILL
        )

        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill   = fill
            cell.border = THIN
            cell.alignment = Alignment(vertical="center")

    _auto_widths(ws)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # ── Style "Summary" sheet + add chart ─────────────────────────────────────
    if "Summary" in wb.sheetnames and cat_col:
        ws2 = wb["Summary"]
        _style_header_row(ws2, summary)

        for row_idx in range(2, ws2.max_row + 1):
            for col_idx in range(1, ws2.max_column + 1):
                cell = ws2.cell(row=row_idx, column=col_idx)
                cell.fill   = SUBTOTAL_FILL
                cell.border = THIN
                cell.alignment = Alignment(vertical="center")

        _auto_widths(ws2)
        ws2.freeze_panes = "A2"

        # Bar chart
        chart = BarChart()
        chart.type       = "col"
        chart.title      = "Budget vs Actual by Category"
        chart.y_axis.title = "Amount"
        chart.x_axis.title = cat_col.title()
        chart.style      = 10
        chart.width      = 20
        chart.height     = 12

        n_rows = ws2.max_row
        budget_ref = Reference(ws2, min_col=2, min_row=1, max_row=n_rows)
        actual_ref = Reference(ws2, min_col=3, min_row=1, max_row=n_rows)
        cats_ref   = Reference(ws2, min_col=1, min_row=2, max_row=n_rows)

        chart.add_data(budget_ref, titles_from_data=True)
        chart.add_data(actual_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        chart.series[0].graphicalProperties.solidFill = "1F4E79"
        chart.series[1].graphicalProperties.solidFill = "ED7D31"

        ws2.add_chart(chart, f"A{n_rows + 3}")

    # ── Legend / key box at top of Variance Report ────────────────────────────
    result = io.BytesIO()
    wb.save(result)
    return result.getvalue()


# ── Helpers ────────────────────────────────────────────────────────────────────

def _detect_columns(df: pd.DataFrame) -> dict:
    """Return {"category": col, "budget": col, "actual": col} via fuzzy matching."""
    lower = {c: c.lower() for c in df.columns}
    result = {}

    for col, lc in lower.items():
        if not result.get("category") and any(k in lc for k in _CATEGORY_ALIASES):
            result["category"] = col
        if not result.get("budget") and any(k in lc for k in _BUDGET_ALIASES):
            result["budget"] = col
        if not result.get("actual") and any(k in lc for k in _ACTUAL_ALIASES):
            result["actual"] = col

    if "budget" not in result or "actual" not in result:
        return {}
    return result


def _style_header_row(ws, df):
    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value     = col_name
        cell.fill      = HEADER_FILL
        cell.font      = HEADER_FONT
        cell.border    = THIN
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 20


def _auto_widths(ws):
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        max_w = 10
        for row_idx in range(1, min(ws.max_row + 1, 200)):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is not None:
                max_w = max(max_w, len(str(val)) + 4)
        ws.column_dimensions[col_letter].width = min(max_w, 40)
